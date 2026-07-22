from openpyxl import Workbook
from openpyxl.styles import Border, Font, Alignment, Side, PatternFill

THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(name="Calibri", bold=True, size=16)
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=13)
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

VEHICLE_ORDER = [
    "Heavy Truck",
    "Medium Truck",
    "Small Truck",
    "Large Bus",
    "Mini Bus",
    "Micro Bus",
    "Utility",
    "Car",
    "Auto Rickshaw",
    "Motor Cycle",
    "Bicycle",
    "Cycle Rickshaw",
    "Cart",
]


class TrafficReportExporter:

    @staticmethod
    def export(counter, filename):
        wb = Workbook()

        for idx, (sheet_name, direction_data) in enumerate([
            ("Incoming", counter.incoming),
            ("Outgoing", counter.outgoing),
        ]):
            if idx == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            _build_sheet(ws, sheet_name, direction_data)

        wb.save(filename)


def _build_sheet(ws, direction_label, counts):
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # ── Row 1: Title ──
    ws.merge_cells("A1:E1")
    title_cell = ws.cell(row=1, column=1, value="Vehicle Traffic Survey Form")
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # ── Row 2: Direction ──
    ws.merge_cells("A2:E2")
    dir_cell = ws.cell(row=2, column=1, value=direction_label)
    dir_cell.font = SUBTITLE_FONT
    dir_cell.alignment = CENTER

    # ── Rows 3-5: Metadata fields ──
    meta = [
        ("Traffic Station No:", ""),
        ("Date:", ""),
        ("Time:", ""),
    ]
    for i, (label, value) in enumerate(meta):
        r = 4 + i
        lbl_cell = ws.cell(row=r, column=1, value=label)
        lbl_cell.font = BOLD_FONT
        lbl_cell.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        val_cell = ws.cell(row=r, column=2, value=value)
        val_cell.font = NORMAL_FONT
        val_cell.alignment = LEFT
        val_cell.border = Border(bottom=THIN)

    # ── Row 8: Table headers ──
    headers = ["Serial No.", "Vehicle Name", "Count", "Checked Data", "Field Data"]
    header_row = 8
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        c.fill = HEADER_FILL

    # ── Data rows ──
    for i, vehicle in enumerate(VEHICLE_ORDER):
        r = header_row + 1 + i
        count = counts.get(vehicle, 0)

        sn = ws.cell(row=r, column=1, value=i + 1)
        sn.alignment = CENTER
        sn.border = BORDER
        sn.font = NORMAL_FONT

        vn = ws.cell(row=r, column=2, value=vehicle)
        vn.alignment = LEFT
        vn.border = BORDER
        vn.font = NORMAL_FONT

        ct = ws.cell(row=r, column=3, value=count)
        ct.alignment = CENTER
        ct.border = BORDER
        ct.font = NORMAL_FONT

        cd = ws.cell(row=r, column=4, value="")
        cd.alignment = CENTER
        cd.border = BORDER
        cd.font = NORMAL_FONT

        fd = ws.cell(row=r, column=5, value="")
        fd.alignment = CENTER
        fd.border = BORDER
        fd.font = NORMAL_FONT

        ws.row_dimensions[r].height = 20

    # ── Total row ──
    total_row = header_row + 1 + len(VEHICLE_ORDER)
    total_val = sum(counts.get(v, 0) for v in VEHICLE_ORDER)

    ws.cell(row=total_row, column=1).border = BORDER

    lbl = ws.cell(row=total_row, column=2, value="Total Traffic")
    lbl.font = BOLD_FONT
    lbl.alignment = LEFT
    lbl.border = BORDER

    tot = ws.cell(row=total_row, column=3, value=total_val)
    tot.font = BOLD_FONT
    tot.alignment = CENTER
    tot.border = BORDER

    ws.cell(row=total_row, column=4).border = BORDER
    ws.cell(row=total_row, column=5).border = BORDER

    ws.row_dimensions[total_row].height = 24

    # ── Column widths ──
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # ── Freeze header ──
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Print area ──
    ws.print_area = f"A1:E{total_row}"
